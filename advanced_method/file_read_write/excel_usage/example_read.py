# @Time : 2020/5/6 20:59 
# @Author : tongyue
from openpyxl import load_workbook

wb = load_workbook(filename = 'empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)
print(sheet_ranges.cell(column=2,row=1).value)

