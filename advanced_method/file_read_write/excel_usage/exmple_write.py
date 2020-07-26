# @Time : 2020/5/6 20:37 
# @Author : tongyue


from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

listData = ['a','b','c']

for row in range(1, 40):
    ws1.append(range(600))
    # ws1.append(listData)

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
     for col in range(27, 54):
         ws3.cell(column=col, row=row, value=get_column_letter(col))
print(ws3['AA10'].value)
wb.save(filename = dest_filename)