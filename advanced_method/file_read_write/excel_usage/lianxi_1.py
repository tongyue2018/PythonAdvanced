# @Time : 2020/5/15 16:05 
# @Time : 2020/5/15 16:05
# @Author : tongyue



from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
file_name = 'lianxi_1.xlsx'

ws1 = wb.create_sheet(title='20180515')
ws1.cell(column=1,row=1,value='第1格')

ws2 = wb.active
ws2.title = 'wode'
ws2.cell(column=2,row=2,value='第2格')

wb.save(filename=file_name)


